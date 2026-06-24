"""
Excel Handler - قالب‌بندی فایل Excel برای import
"""


def generate_import_template() -> bytes:
    """
    تولید فایل Excel قالب برای import محصولات.
    کاربران این قالب را دانلود کرده و پر می‌کنند.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'محصولات'

    # هدرها
    headers = [
        'نام محصول',
        'دسته‌بندی',
        'واحد',
        'بارکد',
        'برند',
        'منشأ',
        'توضیحات',
    ]

    # استایل هدر
    header_fill = PatternFill(
        start_color='366092',
        end_color='366092',
        fill_type='solid'
    )
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # عرض ستون‌ها
    column_widths = [25, 20, 10, 15, 15, 15, 30]
    for col_idx, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    # ردیف نمونه
    sample_data = [
        'مرغ گرم',
        'گوشت مرغ',
        'kg',
        '',
        'مرغ ایرانی',
        'ایران',
        'مرغ تازه و گرم',
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = worksheet.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(
            start_color='EBF1DE',
            end_color='EBF1DE',
            fill_type='solid'
        )

    # شیت راهنما
    guide_sheet = workbook.create_sheet(title='راهنما')
    guide_data = [
        ['ستون', 'توضیحات', 'اجباری'],
        ['نام محصول', 'نام کامل محصول', 'بله'],
        ['دسته‌بندی', 'نام دسته‌بندی (اگر وجود نداشت ایجاد می‌شود)', 'بله'],
        ['واحد', 'نماد واحد اندازه‌گیری (مثال: kg, piece, liter)', 'بله'],
        ['بارکد', 'بارکد استاندارد محصول', 'خیر'],
        ['برند', 'نام برند یا تولیدکننده', 'خیر'],
        ['منشأ', 'کشور یا شهر تولید', 'خیر'],
        ['توضیحات', 'توضیحات تکمیلی محصول', 'خیر'],
    ]
    for row_data in guide_data:
        guide_sheet.append(row_data)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()