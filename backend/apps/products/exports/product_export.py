"""
Product Export - خروجی اکسل محصولات
"""
import logging
from apps.products.selectors import ProductSelector

logger = logging.getLogger(__name__)


def export_products_to_excel(
    category_id: int = None,
    is_active: bool = True
) -> bytes:
    """
    خروجی اکسل لیست محصولات.

    Args:
        category_id: فیلتر بر اساس دسته‌بندی (اختیاری)
        is_active: فقط محصولات فعال

    Returns:
        bytes: محتوای فایل Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    if category_id:
        products = ProductSelector.get_by_category(category_id)
    else:
        products = ProductSelector.get_all_active()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'محصولات'

    # هدرها
    headers = [
        'شناسه',
        'نام محصول',
        'دسته‌بندی',
        'واحد',
        'بارکد',
        'برند',
        'منشأ',
        'فعال',
        'تاریخ ایجاد',
    ]

    header_fill = PatternFill(
        start_color='366092',
        end_color='366092',
        fill_type='solid'
    )
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # داده‌ها
    for row_idx, product in enumerate(products, start=2):
        worksheet.cell(row=row_idx, column=1, value=product.id)
        worksheet.cell(row=row_idx, column=2, value=product.name)
        worksheet.cell(
            row=row_idx,
            column=3,
            value=product.category.name
        )
        worksheet.cell(
            row=row_idx,
            column=4,
            value=f'{product.unit.name} ({product.unit.symbol})'
        )
        worksheet.cell(
            row=row_idx,
            column=5,
            value=product.barcode or ''
        )
        worksheet.cell(row=row_idx, column=6, value=product.brand)
        worksheet.cell(row=row_idx, column=7, value=product.origin)
        worksheet.cell(
            row=row_idx,
            column=8,
            value='بله' if product.is_active else 'خیر'
        )
        worksheet.cell(
            row=row_idx,
            column=9,
            value=product.created_at.strftime('%Y-%m-%d')
        )

    # عرض ستون‌ها
    column_widths = [8, 30, 20, 15, 15, 15, 15, 8, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()